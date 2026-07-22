# -*- coding: utf-8 -*-
"""
将 CarMove 批量截图目录中的轨迹大图和进场小图汇总到 Excel。

文件名格式：
    大图：车牌号_YYYY-MM-DD_YYYY-MM-DD.png
    小图：车牌号_YYYY-MM-DD_YYYY-MM-DD_target.png

示例：
    冀J6C159_2025-12-14_2025-12-14.png
    冀J6C159_2025-12-14_2025-12-14_target.png

运行：
    pip install openpyxl pillow
    # 新建图片汇总 Excel
    python scripts/export_screenshot_images_to_excel.py 图片目录 截图汇总.xlsx

    # 将原图插入已有 Excel，按每行“日期 + 车号”匹配
    python scripts/export_screenshot_images_to_excel.py 图片目录 --workbook 原表.xlsx -o 原表_插图.xlsx
"""
import argparse
import copy
import tempfile
from datetime import date, datetime
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from PIL import Image as PilImage


IMAGE_EXTS = {".png", ".jpg", ".jpeg"}
FILE_RE = re.compile(
    r"^(?P<plate>.+)_(?P<start>\d{4}-\d{2}-\d{2})(?:_|-)"
    r"(?P<end>\d{4}-\d{2}-\d{2})(?P<target>_target)?$"
)


@dataclass(frozen=True)
class ScreenshotKey:
    plate: str
    start_date: str
    end_date: str


@dataclass
class ScreenshotPair:
    key: ScreenshotKey
    trajectory_path: Path | None = None
    target_path: Path | None = None


def parse_image_name(path: Path):
    if path.suffix.lower() not in IMAGE_EXTS:
        return None
    match = FILE_RE.match(path.stem)
    if not match:
        return None
    key = ScreenshotKey(
        plate=match.group("plate"),
        start_date=match.group("start"),
        end_date=match.group("end"),
    )
    return key, bool(match.group("target"))


def collect_screenshots(image_dir: Path, recursive: bool) -> list[ScreenshotPair]:
    files = image_dir.rglob("*") if recursive else image_dir.iterdir()
    pairs: dict[ScreenshotKey, ScreenshotPair] = {}

    for path in files:
        if not path.is_file():
            continue
        parsed = parse_image_name(path)
        if not parsed:
            continue
        key, is_target = parsed
        pair = pairs.setdefault(key, ScreenshotPair(key=key))
        if is_target:
            pair.target_path = path
        else:
            pair.trajectory_path = path

    return sorted(pairs.values(), key=screenshot_created_time)


def screenshot_created_time(pair: ScreenshotPair) -> tuple[float, str, str, str]:
    image_path = pair.trajectory_path or pair.target_path
    created_time = image_path.stat().st_ctime if image_path else 0
    return (created_time, pair.key.plate, pair.key.start_date, pair.key.end_date)


def fit_size(path: Path, max_width: int, max_height: int) -> tuple[int, int]:
    with PilImage.open(path) as img:
        width, height = img.size
    if width <= 0 or height <= 0:
        return max_width, max_height
    scale = min(max_width / width, max_height / height, 1.0)
    return int(width * scale), int(height * scale)


class ImageEmbedder:
    def __init__(self, temp_dir: Path, quality: int = 85):
        self.temp_dir = temp_dir
        self.quality = quality
        self.cache: dict[tuple[Path, int, int], tuple[Path, int, int]] = {}

    def prepare(self, path: Path, max_width: int, max_height: int) -> tuple[Path, int, int]:
        key = (path, max_width, max_height)
        if key in self.cache:
            return self.cache[key]

        with PilImage.open(path) as img:
            width, height = img.size
            if width <= 0 or height <= 0:
                size = (max_width, max_height)
            else:
                scale = min(max_width / width, max_height / height, 1.0)
                size = (max(1, int(width * scale)), max(1, int(height * scale)))

            resized = img.convert("RGB")
            if resized.size != size:
                resized = resized.resize(size, PilImage.Resampling.LANCZOS)

            out_path = self.temp_dir / f"img_{len(self.cache) + 1}.jpg"
            resized.save(out_path, "JPEG", quality=self.quality, optimize=True)

        result = (out_path, size[0], size[1])
        self.cache[key] = result
        return result


def add_image(ws, path: Path | None, cell: str, max_width: int, max_height: int,
              embedder: ImageEmbedder | None = None):
    if not path or not path.exists():
        ws[cell] = "未找到图片"
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        return
    if embedder:
        image_path, width, height = embedder.prepare(path, max_width, max_height)
    else:
        image_path = path
        width, height = fit_size(path, max_width, max_height)
    image = XlsxImage(str(image_path))
    image.width, image.height = width, height
    image.anchor = cell
    ws.add_image(image)


def image_index(pairs: list[ScreenshotPair]) -> dict[ScreenshotKey, ScreenshotPair]:
    return {pair.key: pair for pair in pairs}


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def fixed_year_from_number_format(number_format: str) -> int | None:
    match = re.search(r"(?:\\\d){4}", number_format or "")
    if not match:
        return None
    return int(match.group(0).replace("\\", ""))


def normalize_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if not text:
        return ""
    match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text)
    if match:
        year, month, day = map(int, match.groups())
        return f"{year:04d}-{month:02d}-{day:02d}"
    match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        return f"{year:04d}-{month:02d}-{day:02d}"
    return ""


def normalize_date_cell(cell) -> str:
    value = cell.value
    fixed_year = fixed_year_from_number_format(cell.number_format)
    if fixed_year and isinstance(value, (datetime, date)):
        return f"{fixed_year:04d}-{value.month:02d}-{value.day:02d}"
    return normalize_date(value)


def find_header_columns(ws) -> tuple[int, int, int] | None:
    for row in range(1, min(ws.max_row, 10) + 1):
        scan_cols = min(ws.max_column, 50)
        headers = {clean_text(ws.cell(row=row, column=col).value): col for col in range(1, scan_cols + 1)}
        date_col = headers.get("过毛时间") or headers.get("毛重时间") or headers.get("日期")
        plate_col = headers.get("车号") or headers.get("车牌号") or headers.get("车牌号码")
        if date_col and plate_col:
            return row, date_col, plate_col
    return None


def actual_content_max_column(ws, scan_cols: int = 50) -> int:
    max_col = 1
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=min(ws.max_column, scan_cols)):
        for cell in row:
            if cell.value not in (None, ""):
                max_col = max(max_col, cell.column)
    return max_col


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy.copy(source._style)
    if source.font:
        target.font = copy.copy(source.font)
    if source.fill:
        target.fill = copy.copy(source.fill)
    if source.border:
        target.border = copy.copy(source.border)
    if source.alignment:
        target.alignment = copy.copy(source.alignment)
    if source.number_format:
        target.number_format = source.number_format


def ensure_image_columns(ws, header_row: int, plate_col: int) -> tuple[int, int]:
    scan_cols = min(ws.max_column, 50)
    existing = {clean_text(ws.cell(row=header_row, column=col).value): col for col in range(1, scan_cols + 1)}
    trajectory_col = existing.get("轨迹图")
    target_col = existing.get("进场图")
    if trajectory_col and target_col:
        return trajectory_col, target_col

    trajectory_col = max(plate_col, actual_content_max_column(ws)) + 1
    target_col = plate_col + 2
    target_col = trajectory_col + 1

    for col, title in ((trajectory_col, "轨迹图"), (target_col, "进场图")):
        source = ws.cell(row=header_row, column=plate_col)
        cell = ws.cell(row=header_row, column=col, value=title)
        copy_cell_style(source, cell)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    return trajectory_col, target_col


def extend_title_merge(ws, last_col: int):
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row == 1 and max_row == 1 and min_col == 1 and max_col < last_col:
            ws.unmerge_cells(str(merged_range))
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
            break


def insert_images_into_workbook(
    workbook_path: Path,
    output_path: Path,
    pairs: list[ScreenshotPair],
    image_width: int,
    trajectory_height: int,
    target_height: int,
    compress_images: bool = False,
) -> tuple[int, int]:
    wb = load_workbook(workbook_path)
    pairs_by_key = image_index(pairs)
    matched_rows = 0
    missing_images = 0

    def process(embedder: ImageEmbedder | None):
        nonlocal matched_rows, missing_images
        for ws in wb.worksheets:
            header = find_header_columns(ws)
            if not header:
                continue
            header_row, date_col, plate_col = header
            trajectory_col, target_col = ensure_image_columns(ws, header_row, plate_col)
            extend_title_merge(ws, max(trajectory_col, target_col))
            ws.column_dimensions[get_column_letter(trajectory_col)].width = max(24, image_width / 7)
            ws.column_dimensions[get_column_letter(target_col)].width = max(24, image_width / 7)

            row_height = max(trajectory_height, target_height) * 0.75
            for row in range(header_row + 1, ws.max_row + 1):
                plate = clean_text(ws.cell(row=row, column=plate_col).value)
                row_date = normalize_date_cell(ws.cell(row=row, column=date_col))
                if not plate or not row_date:
                    continue

                key = ScreenshotKey(plate=plate, start_date=row_date, end_date=row_date)
                pair = pairs_by_key.get(key)
                if not pair:
                    missing_images += 1
                    continue

                ws.row_dimensions[row].height = row_height
                for col in (trajectory_col, target_col):
                    copy_cell_style(ws.cell(row=row, column=plate_col), ws.cell(row=row, column=col))
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

                add_image(ws, pair.trajectory_path, f"{get_column_letter(trajectory_col)}{row}",
                          image_width, trajectory_height, embedder)
                add_image(ws, pair.target_path, f"{get_column_letter(target_col)}{row}",
                          image_width, target_height, embedder)
                matched_rows += 1

    if compress_images:
        with tempfile.TemporaryDirectory(prefix="carmove_excel_images_") as temp_dir:
            process(ImageEmbedder(Path(temp_dir)))
    else:
        process(None)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return matched_rows, missing_images


def write_excel(
    pairs: list[ScreenshotPair],
    output_path: Path,
    image_width: int,
    trajectory_height: int,
    target_height: int,
    compress_images: bool = False,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "截图汇总"

    headers = ["车牌号", "开始时间", "结束时间", "轨迹图", "进场图"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = [18, 16, 16, max(24, image_width / 7), max(24, image_width / 7)]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    row_height = max(trajectory_height, target_height) * 0.75
    def process(embedder: ImageEmbedder | None):
        for row, pair in enumerate(pairs, start=2):
            ws.row_dimensions[row].height = row_height
            values = [pair.key.plate, pair.key.start_date, pair.key.end_date]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border

            add_image(ws, pair.trajectory_path, f"D{row}", image_width, trajectory_height, embedder)
            add_image(ws, pair.target_path, f"E{row}", image_width, target_height, embedder)

    if compress_images:
        with tempfile.TemporaryDirectory(prefix="carmove_excel_images_") as temp_dir:
            process(ImageEmbedder(Path(temp_dir)))
    else:
        process(None)

    ws.freeze_panes = "A2"

    summary = wb.create_sheet("说明")
    summary.append(["项目", "内容"])
    summary.append(["记录数", len(pairs)])
    summary.append(["说明", "每一行对应一组车牌号、开始时间、结束时间、轨迹图、进场图。"])
    summary.append(["缺图", "如果没有找到对应图片，单元格会显示“未找到图片”。"])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 80

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="将 CarMove 截图图片汇总到 Excel，或插入已有 Excel。")
    parser.add_argument("image_dir", help="截图图片所在目录")
    parser.add_argument("output_positional", nargs="?", help="输出 Excel 文件路径")
    parser.add_argument("-o", "--output", default="截图汇总.xlsx", help="输出 Excel 文件路径")
    parser.add_argument("--workbook", help="已有 Excel 文件路径；提供后按每行日期和车号插图")
    parser.add_argument("-r", "--recursive", action="store_true", help="递归扫描子目录")
    parser.add_argument("--image-width", type=int, default=360, help="Excel 中图片最大宽度（像素）")
    parser.add_argument("--trajectory-height", type=int, default=220, help="轨迹图最大高度（像素）")
    parser.add_argument("--target-height", type=int, default=220, help="进场图最大高度（像素）")
    parser.add_argument("--compress-images", action="store_true", help="压缩图片后再嵌入；默认嵌入原图")
    args = parser.parse_args()

    image_dir = Path(args.image_dir).expanduser().resolve()
    workbook_path = Path(args.workbook).expanduser().resolve() if args.workbook else None
    if workbook_path and args.output == "截图汇总.xlsx" and not args.output_positional:
        output_path = workbook_path.with_name(f"{workbook_path.stem}_插图{workbook_path.suffix}")
    else:
        output_path = Path(args.output_positional or args.output).expanduser().resolve()
    if not image_dir.exists() or not image_dir.is_dir():
        raise SystemExit(f"图片目录不存在：{image_dir}")

    pairs = collect_screenshots(image_dir, args.recursive)
    if not pairs:
        raise SystemExit("没有找到符合命名规则的截图图片。")

    if workbook_path:
        if not workbook_path.exists() or not workbook_path.is_file():
            raise SystemExit(f"Excel 文件不存在：{workbook_path}")
        matched_rows, missing_images = insert_images_into_workbook(
            workbook_path=workbook_path,
            output_path=output_path,
            pairs=pairs,
            image_width=args.image_width,
            trajectory_height=args.trajectory_height,
            target_height=args.target_height,
            compress_images=args.compress_images,
        )
        print(f"完成：{output_path}，插入 {matched_rows} 行，未匹配 {missing_images} 行")
    else:
        write_excel(
            pairs=pairs,
            output_path=output_path,
            image_width=args.image_width,
            trajectory_height=args.trajectory_height,
            target_height=args.target_height,
            compress_images=args.compress_images,
        )
        print(f"完成：{output_path}，共 {len(pairs)} 组")


if __name__ == "__main__":
    main()
